"""
Formula-linked multi-sheet DCF workbook (Option B).
Sheets: Assumptions, DCF, Working Capital, Debt, Summary, Sensitivity.
Inputs seeded from valuation JSON; calculation cells use Excel formulas.
"""

from __future__ import annotations

import logging
import math
from io import BytesIO
from typing import Any

logger = logging.getLogger("valuation-api")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WorkbookType


def _num(x: Any, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        return float(x)
    except (TypeError, ValueError):
        return default


def _scenario_triplet(data: dict[str, Any], key: str, default: tuple[float, float, float]) -> tuple[float, float, float]:
    scenarios = data.get("scenarios") or {}
    bear = _num((scenarios.get("bear") or {}).get(key), default[0])
    base = _num((scenarios.get("base") or {}).get(key), default[1])
    bull = _num((scenarios.get("bull") or {}).get(key), default[2])
    if bear == 0 and base == 0 and bull == 0:
        return default
    return bear, base, bull


def _ltm_revenue_millions(data: dict[str, Any], base_growth: float) -> float:
    hist = data.get("historical") or {}
    revs = hist.get("revenue") or []
    if revs:
        return _num(revs[-1])
    proj = (data.get("projections") or {}).get("base") or {}
    arr = proj.get("revenue") or []
    if not arr:
        return 100_000.0
    y0 = _num(arr[0])
    g = base_growth if abs(base_growth) > 1e-6 else 0.08
    if abs(base_growth) <= 1:
        g = base_growth
    return y0 / (1.0 + g)


def _shares_millions(data: dict[str, Any]) -> float:
    cap_bn = _num(data.get("market_cap_bn"))
    px = _num(data.get("current_price"))
    if px > 0 and cap_bn > 0:
        return cap_bn * 1000.0 / px
    return 1500.0


def _net_debt_millions(data: dict[str, Any]) -> float:
    ev_bn = _num(data.get("enterprise_value_bn"))
    cap_bn = _num(data.get("market_cap_bn"))
    if ev_bn > 0 and cap_bn > 0:
        return (ev_bn - cap_bn) * 1000.0
    return 50_000.0


def _pick(bear: float, base: float, bull: float, scenario: int = 2) -> float:
    return (bear, base, bull)[max(0, min(scenario - 1, 2))]


def _compute_dcf_cache(
    *,
    ltm: float,
    rev_g: float,
    tg: float,
    ebitda_m: float,
    nwc_pct_rev: float,
    capex_pct: float,
    da_pct: float,
    tax: float,
    wacc: float,
    nd: float,
    shares: float,
    exit_mult: float,
    terminal_method: int,
) -> dict[tuple[int, int], float]:
    """Mirror DCF formulas in Python so Excel can show cached results on open."""
    revs: list[float] = []
    ebitdas: list[float] = []
    das: list[float] = []
    ebits: list[float] = []
    taxes: list[float] = []
    nopats: list[float] = []
    capexs: list[float] = []
    dnwc: list[float] = []
    fcfs: list[float] = []

    prev_rev = ltm
    for _ in range(5):
        rev = prev_rev * (1 + rev_g)
        revs.append(rev)
        ebitda = rev * ebitda_m
        da = rev * da_pct
        ebit = ebitda - da
        tax_amt = ebit * tax
        nopat = ebit - tax_amt
        capex = -rev * capex_pct
        delta = -(rev - prev_rev) * nwc_pct_rev
        prev_rev = rev
        fcf = nopat + da + capex + delta
        ebitdas.append(ebitda)
        das.append(da)
        ebits.append(ebit)
        taxes.append(tax_amt)
        nopats.append(nopat)
        capexs.append(capex)
        dnwc.append(delta)
        fcfs.append(fcf)

    last_fcf = fcfs[-1]
    last_ebitda = ebitdas[-1]
    tv_gordon = last_fcf * (1 + tg) / (wacc - tg) if wacc > tg else 0.0
    tv_exit = last_ebitda * exit_mult
    tv = tv_gordon if terminal_method == 1 else tv_exit

    pv_fcf = sum(fcf / ((1 + wacc) ** (i + 1)) for i, fcf in enumerate(fcfs))
    pv_tv = tv / ((1 + wacc) ** 5)
    ev = pv_fcf + pv_tv
    equity = ev - nd
    price = equity / shares if shares else 0.0

    cache: dict[tuple[int, int], float] = {}
    for i in range(5):
        col = 3 + i
        cache[(5, col)] = revs[i]
        cache[(7, col)] = ebitdas[i]
        cache[(8, col)] = das[i]
        cache[(9, col)] = ebits[i]
        cache[(10, col)] = taxes[i]
        cache[(11, col)] = nopats[i]
        cache[(12, col)] = capexs[i]
        cache[(13, col)] = dnwc[i]
        cache[(14, col)] = fcfs[i]
        cache[(20, col)] = fcfs[i] / ((1 + wacc) ** (i + 1))

    cache[(16, 7)] = tv_gordon
    cache[(17, 7)] = tv_exit
    cache[(18, 7)] = tv
    cache[(21, 7)] = pv_tv
    cache[(23, 2)] = ev
    cache[(24, 2)] = equity
    cache[(25, 2)] = price
    return cache


def _apply_formula_cache(ws, cache: dict[tuple[int, int], float]) -> None:
    """Attach cached numeric results to formula cells (Excel shows values before F9)."""
    from openpyxl.xml.constants import SHEET_MAIN_NS
    from openpyxl.xml.functions import SubElement

    for (row, col), result in cache.items():
        try:
            if not math.isfinite(result):
                result = 0.0
            cell = ws.cell(row=row, column=col)
            formula = cell.value
            if not formula or not isinstance(formula, str) or not formula.startswith("="):
                continue
            el = cell._element
            v_el = el.find(f"{{{SHEET_MAIN_NS}}}v")
            if v_el is None:
                v_el = SubElement(el, f"{{{SHEET_MAIN_NS}}}v")
            v_el.text = str(float(result))
            cell.data_type = "f"
        except Exception as exc:
            logger.debug("Skip cache for %s!%s%s: %s", ws.title, get_column_letter(col), row, exc)


def build_formula_linked_workbook(data: dict[str, Any], ticker: str) -> BytesIO:
    wb = Workbook()
    # Remove default sheet after we create ordered sheets
    default = wb.active
    wb.remove(default)

    ws_a = wb.create_sheet("Assumptions", 0)
    ws_dcf = wb.create_sheet("DCF", 1)
    ws_wc = wb.create_sheet("Working Capital", 2)
    ws_debt = wb.create_sheet("Debt", 3)
    ws_sum = wb.create_sheet("Summary", 4)
    ws_sens = wb.create_sheet("Sensitivity", 5)

    wacc = data.get("wacc") or {}
    rf = _num(wacc.get("risk_free_rate"), 0.042)
    if rf > 1:
        rf = rf / 100.0
    beta = _num(wacc.get("beta"), 1.1)
    erp = _num(wacc.get("equity_risk_premium"), 0.055)
    if erp > 1:
        erp = erp / 100.0
    tax = _num(wacc.get("tax_rate"), 0.21)
    if tax > 1:
        tax = tax / 100.0
    kd = _num(wacc.get("cost_of_debt"), 0.045)
    if kd > 1:
        kd = kd / 100.0
    wd = _num(wacc.get("debt_weight"), 0.15)
    we = _num(wacc.get("equity_weight"), 0.85)

    rev_g_bear, rev_g_base, rev_g_bull = _scenario_triplet(
        data, "revenue_cagr", (0.05, 0.10, 0.15)
    )
    tg_bear, tg_base, tg_bull = _scenario_triplet(
        data, "terminal_growth", (0.02, 0.025, 0.03)
    )
    if abs(rev_g_base) > 1:
        rev_g_bear, rev_g_base, rev_g_bull = rev_g_bear / 100, rev_g_base / 100, rev_g_bull / 100
    if abs(tg_base) > 1:
        tg_bear, tg_base, tg_bull = tg_bear / 100, tg_base / 100, tg_bull / 100

    proj = (data.get("projections") or {}).get("base") or {}
    years = proj.get("years") or ["Y1", "Y2", "Y3", "Y4", "Y5"]
    revs = proj.get("revenue") or []
    ebitda_m = 0.32
    if revs and proj.get("ebitda"):
        ebitda_m = _num(proj["ebitda"][0]) / _num(revs[0]) if _num(revs[0]) else 0.32
    em_bear, em_base, em_bull = _scenario_triplet(
        data, "exit_multiple", (10.0, 12.0, 14.0)
    )

    input_fill = PatternFill("solid", fgColor="B4C7E7")
    input_font = Font(color="1F2937", bold=False)
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(color="FFFFFF", bold=True)

    # ----- Assumptions -----
    ws_a["A1"] = f"{data.get('company_name', ticker)} — model drivers"
    ws_a["A1"].font = Font(bold=True, size=14)

    ws_a["A3"] = "Scenario (1=Bear, 2=Base, 3=Bull)"
    ws_a["B4"] = 2
    ws_a["B4"].fill = input_fill
    ws_a["B4"].font = input_font

    ws_a["A5"] = "Driver"
    ws_a["B5"] = "Bear"
    ws_a["C5"] = "Base"
    ws_a["D5"] = "Bull"
    ws_a["E5"] = "Active (formula)"
    for c in range(1, 6):
        cell = ws_a.cell(5, c)
        cell.fill = hdr_fill
        cell.font = hdr_font

    drivers = [
        ("Revenue YoY growth", rev_g_bear, rev_g_base, rev_g_bull, 6),
        ("Terminal growth (g)", tg_bear, tg_base, tg_bull, 7),
        ("EBITDA margin", 0.28, ebitda_m, 0.36, 8),
        ("ΔNWC as % of ΔRevenue", 0.02, 0.025, 0.03, 9),
        ("CapEx as % of Revenue", 0.045, 0.05, 0.055, 10),
        ("D&A as % of Revenue", 0.04, 0.05, 0.06, 11),
        ("Exit EV/EBITDA (terminal check)", em_bear, em_base, em_bull, 12),
    ]
    for label, vb, vbase, vbull, row in drivers:
        ws_a.cell(row, 1, label)
        ws_a.cell(row, 2, vb)
        ws_a.cell(row, 3, vbase)
        ws_a.cell(row, 4, vbull)
        for col in (2, 3, 4):
            ws_a.cell(row, col).fill = input_fill
            ws_a.cell(row, col).font = input_font
        ws_a.cell(row, 5, f"=CHOOSE($B$4,B{row},C{row},D{row})")

    # WACC build (inputs + formulas) rows 15-24
    ws_a["A15"] = "WACC inputs"
    ws_a["A15"].font = Font(bold=True)
    wacc_rows = [
        (16, "Risk-free rate", rf),
        (17, "Beta", beta),
        (18, "Equity risk premium", erp),
        (19, "Tax rate", tax),
        (20, "Pre-tax cost of debt", kd),
        (21, "Debt weight (Wd)", wd),
        (22, "Equity weight (We)", we),
    ]
    for row, label, val in wacc_rows:
        ws_a.cell(row, 1, label)
        ws_a.cell(row, 2, val)
        ws_a.cell(row, 2).fill = input_fill
        ws_a.cell(row, 2).font = input_font

    ws_a["A23"] = "Cost of equity (CAPM)"
    ws_a["B23"] = "=B16+B17*B18"
    ws_a["A24"] = "After-tax cost of debt"
    ws_a["B24"] = "=B20*(1-B19)"
    ws_a["A25"] = "WACC"
    ws_a["B25"] = "=B22*B23+B21*B24"

    ltm = _ltm_revenue_millions(data, rev_g_base)
    nd = _net_debt_millions(data)
    sh = _shares_millions(data)

    ws_a["A27"] = "Balance bridge inputs ($M except shares)"
    ws_a["A28"] = "LTM Revenue (anchor)"
    ws_a["B28"] = ltm
    ws_a["B28"].fill = input_fill
    ws_a["A29"] = "Net debt (manual override OK)"
    ws_a["B29"] = nd
    ws_a["B29"].fill = input_fill
    ws_a["A30"] = "Diluted shares (M)"
    ws_a["B30"] = sh
    ws_a["B30"].fill = input_fill

    ws_a["A32"] = "Terminal method (1=Gordon Growth, 2=Exit EV/EBITDA)"
    ws_a["B32"] = 2
    ws_a["B32"].fill = input_fill
    ws_a["B32"].font = input_font

    # ----- DCF -----
    ws_dcf["A1"] = "Discounted cash flow (formulas → Assumptions)"
    ws_dcf["A1"].font = Font(bold=True, size=14)
    ws_dcf["A2"] = (
        "Projections in columns C–G. If blank: click Enable Editing (yellow bar), then press F9. "
        "Values are pre-cached; formulas remain editable."
    )
    ws_dcf["A2"].font = Font(italic=True, size=10, color="666666")

    ws_dcf["B3"] = "Year index"
    for i, y in enumerate(years[:5]):
        col = 3 + i
        ws_dcf.cell(3, col, i + 1)
        ws_dcf.cell(4, col, str(y))
        ws_dcf.cell(4, col).fill = hdr_fill
        ws_dcf.cell(4, col).font = hdr_font

    ws_dcf["A5"] = "Revenue"
    for i in range(5):
        col = 3 + i
        col_letter = get_column_letter(col)
        prev = get_column_letter(col - 1) if i > 0 else None
        if i == 0:
            ws_dcf.cell(5, col, "=Assumptions!$B$28*(1+Assumptions!$E$6)")
        else:
            ws_dcf.cell(5, col, f"={prev}5*(1+Assumptions!$E$6)")

    ws_dcf["A7"] = "EBITDA"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_dcf.cell(7, col, f"={cl}5*Assumptions!$E$8")

    ws_dcf["A8"] = "D&A"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_dcf.cell(8, col, f"={cl}5*Assumptions!$E$11")

    ws_dcf["A9"] = "EBIT"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_dcf.cell(9, col, f"={cl}7-{cl}8")

    ws_dcf["A10"] = "Taxes"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_dcf.cell(10, col, f"={cl}9*Assumptions!$B$19")

    ws_dcf["A11"] = "NOPAT"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_dcf.cell(11, col, f"={cl}9-{cl}10")

    ws_dcf["A12"] = "CapEx"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_dcf.cell(12, col, f"=-{cl}5*Assumptions!$E$10")

    # ΔNWC on DCF (avoids circular ref with Working Capital sheet)
    ws_dcf["A13"] = "Δ Net working capital"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        if i == 0:
            ws_dcf.cell(13, col, f"=-({cl}5-Assumptions!$B$28)*Assumptions!$E$9")
        else:
            pcl = get_column_letter(col - 1)
            ws_dcf.cell(13, col, f"=-({cl}5-{pcl}5)*Assumptions!$E$9")

    ws_dcf["A14"] = "Unlevered FCF"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_dcf.cell(14, col, f"={cl}11+{cl}8+{cl}12+{cl}13")

    ws_dcf["A16"] = "Terminal value — Gordon Growth"
    g_col = get_column_letter(3 + 4)
    ws_dcf.cell(16, 7, f"={g_col}14*(1+Assumptions!$E$7)/(Assumptions!$B$25-Assumptions!$E$7)")

    ws_dcf["A17"] = "Terminal value — Exit EV/EBITDA"
    ws_dcf.cell(17, 7, f"={g_col}7*Assumptions!$E$12")

    ws_dcf["A18"] = "Selected terminal value"
    ws_dcf.cell(18, 7, "=IF(Assumptions!$B$32=1,G16,G17)")

    ws_dcf["A20"] = "PV of UFCF"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_dcf.cell(20, col, f"={cl}14/(1+Assumptions!$B$25)^{cl}3")

    ws_dcf["A21"] = "PV of terminal"
    ws_dcf.cell(21, 7, f"=G18/(1+Assumptions!$B$25)^G3")

    ws_dcf["A23"] = "Enterprise value"
    ws_dcf["B23"] = "=SUM(C20:G20)+G21"

    ws_dcf["A24"] = "Equity value"
    ws_dcf["B24"] = "=B23-Assumptions!$B$29"

    ws_dcf["A25"] = "Implied price / share"
    ws_dcf["B25"] = "=B24/Assumptions!$B$30"

    # ----- Working Capital (NWC level row 5, ΔNWC row 6 → DCF row 13) -----
    ws_wc["A1"] = "Net working capital (linked to DCF revenue)"
    ws_wc["A1"].font = Font(bold=True, size=14)
    ws_wc["A3"] = "NWC % of revenue (= Assumptions active driver)"
    ws_wc["B3"] = "=Assumptions!$E$9"
    ws_wc["A4"] = "Year"
    for i, y in enumerate(years[:5]):
        col = 3 + i
        ws_wc.cell(4, col, str(y))
        ws_wc.cell(4, col).fill = hdr_fill
        ws_wc.cell(4, col).font = hdr_font

    ws_wc["A5"] = "NWC level ($M)"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_wc.cell(5, col, f"=DCF!{cl}5*$B$3")

    ws_wc["A6"] = "Δ NWC ($M, mirrors DCF row 13)"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_wc.cell(6, col, f"=DCF!{cl}13")

    # ----- Debt -----
    ws_debt["A1"] = "Debt schedule (simplified bullet)"
    ws_debt["A1"].font = Font(bold=True, size=14)
    ws_debt["A3"] = "Opening net debt (from Assumptions)"
    ws_debt["B3"] = "=Assumptions!$B$29"
    ws_debt["A4"] = "Interest rate"
    ws_debt["B4"] = "=Assumptions!$B$20"
    ws_debt["A5"] = "Year"
    for i, y in enumerate(years[:5]):
        col = 3 + i
        ws_debt.cell(5, col, str(y))
        ws_debt.cell(5, col).fill = hdr_fill
        ws_debt.cell(5, col).font = hdr_font

    ws_debt["A6"] = "Beg. debt balance"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        if i == 0:
            ws_debt.cell(6, col, "=$B$3")
        else:
            pcl = get_column_letter(col - 1)
            ws_debt.cell(6, col, f"={pcl}9")

    ws_debt["A7"] = "Interest expense"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_debt.cell(7, col, f"={cl}6*$B$4")

    ws_debt["A8"] = "Mandatory repayment"
    for i in range(5):
        col = 3 + i
        ws_debt.cell(8, col, 0)

    ws_debt["A9"] = "Ending debt balance"
    for i in range(5):
        col = 3 + i
        cl = get_column_letter(col)
        ws_debt.cell(9, col, f"={cl}6-{cl}8")

    # ----- Summary (values from JSON + links to model) -----
    ws_sum["A1"] = f"{ticker} — valuation summary"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A3"] = "Scenario outputs (from API / Claude JSON)"
    ws_sum["A4"] = "Scenario"
    ws_sum["B4"] = "Implied price"
    ws_sum["C4"] = "Upside %"
    for c in range(1, 4):
        ws_sum.cell(4, c).fill = hdr_fill
        ws_sum.cell(4, c).font = hdr_font
    row = 4
    scenarios = data.get("scenarios") or {}
    for name in ("bear", "base", "bull"):
        s = scenarios.get(name) or {}
        row += 1
        ws_sum.cell(row, 1, name.title())
        ws_sum.cell(row, 2, _num(s.get("implied_price")))
        ws_sum.cell(row, 3, _num(s.get("upside_pct")))
    ws_sum.cell(row + 2, 1, "Model-implied price (Base drivers, DCF sheet)")
    ws_sum.cell(row + 2, 2, "=DCF!$B$25")
    ws_sum.cell(row + 3, 1, "Model enterprise value ($M)")
    ws_sum.cell(row + 3, 2, "=DCF!$B$23")
    ws_sum.cell(row + 4, 1, "Terminal method in use")
    ws_sum.cell(row + 4, 2, '=IF(Assumptions!$B$32=1,"Gordon Growth","Exit EV/EBITDA")')

    _build_sensitivity_sheet(ws_sens, hdr_fill, hdr_font, input_fill)

    ke = rf + beta * erp
    kd_at = kd * (1 - tax)
    wacc_val = we * ke + wd * kd_at
    if _num(wacc.get("wacc")) > 0:
        wacc_val = _num(wacc.get("wacc"))
        if wacc_val > 1:
            wacc_val /= 100.0

    active_rev_g = _pick(rev_g_bear, rev_g_base, rev_g_bull, 2)
    active_tg = _pick(tg_bear, tg_base, tg_bull, 2)
    active_ebitda_m = _pick(0.28, ebitda_m, 0.36, 2)
    active_nwc = _pick(0.02, 0.025, 0.03, 2)
    active_capex = _pick(0.045, 0.05, 0.055, 2)
    active_da = _pick(0.04, 0.05, 0.06, 2)

    dcf_cache = _compute_dcf_cache(
        ltm=ltm,
        rev_g=active_rev_g,
        tg=active_tg,
        ebitda_m=active_ebitda_m,
        nwc_pct_rev=active_nwc,
        capex_pct=active_capex,
        da_pct=active_da,
        tax=tax,
        wacc=wacc_val,
        nd=nd,
        shares=sh,
        exit_mult=em_base,
        terminal_method=2,
    )
    try:
        _apply_formula_cache(ws_dcf, dcf_cache)
    except Exception as exc:
        logger.warning("DCF formula cache skipped: %s", exc)

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception as exc:
        logger.warning("Workbook calc properties skipped: %s", exc)

    _autosize_columns(wb)
    buf = BytesIO()
    try:
        wb.save(buf)
    except Exception as exc:
        logger.exception("Excel save failed")
        raise RuntimeError(f"Could not build Excel file: {exc}") from exc
    buf.seek(0)
    return buf


def _build_sensitivity_sheet(ws, hdr_fill: PatternFill, hdr_font: Font, input_fill: PatternFill) -> None:
    """5×5 implied price table: WACC (rows) vs terminal g (columns). Center = base case."""
    center_fill = PatternFill("solid", fgColor="BDD7EE")

    ws["A1"] = "Sensitivity — implied price / share ($)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Full DCF recalc per cell; uses same terminal switch as Assumptions!B32"
    ws["A4"] = "WACC ↓ / Terminal g →"

    # Column headers (terminal growth) — row 4, cols B–F
    ws["D4"] = "=Assumptions!$E$7"
    ws["B4"] = "=D4-0.005"
    ws["C4"] = "=D4-0.0025"
    ws["E4"] = "=D4+0.0025"
    ws["F4"] = "=D4+0.005"
    for col in range(2, 7):
        c = ws.cell(4, col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.number_format = "0.00%"

    # Row headers (WACC) — col A, rows 5–9
    ws["A7"] = "=Assumptions!$B$25"
    ws["A5"] = "=A7-0.01"
    ws["A6"] = "=A7-0.005"
    ws["A8"] = "=A7+0.005"
    ws["A9"] = "=A7+0.01"
    for row in range(5, 10):
        c = ws.cell(row, 1)
        c.font = hdr_font
        c.fill = hdr_fill
        c.number_format = "0.00%"

    year_cols = [get_column_letter(3 + i) for i in range(5)]

    for row in range(5, 10):
        fcf_pv = "+".join(
            f"DCF!{yc}14/(1+$A{row})^DCF!{yc}$3" for yc in year_cols
        )
        for col in range(2, 7):
            cl = get_column_letter(col)
            g_ref = f"{cl}$4"
            wacc_ref = f"$A{row}"
            tv_gordon = f"DCF!G14*(1+{g_ref})/({wacc_ref}-{g_ref})"
            tv_exit = "DCF!G7*Assumptions!$E$12"
            tv = f"IF(Assumptions!$B$32=1,{tv_gordon},{tv_exit})"
            pv_tv = f"({tv})/(1+{wacc_ref})^DCF!G$3"
            formula = (
                f"=IF({wacc_ref}<={g_ref},\"—\","
                f"(({fcf_pv})+{pv_tv}-Assumptions!$B$29)/Assumptions!$B$30)"
            )
            cell = ws.cell(row, col, formula)
            cell.number_format = "$#,##0.00"
            if row == 7 and col == 4:
                cell.fill = center_fill
                cell.font = Font(bold=True)


def _autosize_columns(wb: WorkbookType, max_width: int = 42) -> None:
    for ws in wb.worksheets:
        for col_idx in range(1, (ws.max_column or 1) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, (ws.max_row or 1) + 1):
                v = ws.cell(row_idx, col_idx).value
                if v is not None:
                    max_len = max(max_len, min(len(str(v)), 60))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)
