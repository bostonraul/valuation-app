import json
import logging
import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx
from anthropic import Anthropic
from dotenv import load_dotenv
from fastapi import Body, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import Client, create_client

logger = logging.getLogger("valuation-api")

load_dotenv(Path(__file__).resolve().parent.parent / ".env")


def _env(name: str, default: str = "") -> str:
    return os.getenv(name, default).strip()


ANTHROPIC_API_KEY = _env("ANTHROPIC_API_KEY")
FMP_API_KEY = _env("FMP_API_KEY")
SUPABASE_URL = _env("SUPABASE_URL")
SUPABASE_ANON_KEY = _env("SUPABASE_ANON_KEY")
FRONTEND_URL = _env("FRONTEND_URL", "http://localhost:3000")

ROOT = Path(__file__).resolve().parent.parent
SKILL_PATH = ROOT / "skills" / "dcf-model" / "SKILL.md"
AGENT_PATH = ROOT / "agents" / "valuation-agent.md"
FMP_BASE = "https://financialmodelingprep.com/api/v3"

def _cors_origins() -> list[str]:
    origins = [FRONTEND_URL, "http://localhost:3000"]
    extra = _env("CORS_ALLOW_ORIGINS")
    if extra:
        origins.extend(o for o in extra.split(",") if o.strip())
    # De-dupe while preserving order
    seen: set[str] = set()
    unique: list[str] = []
    for o in origins:
        if o and o not in seen:
            seen.add(o)
            unique.append(o)
    return unique


app = FastAPI(title="Valuation API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins(),
    # Matches production + preview Vercel URLs (e.g. valuation-app-phi.vercel.app)
    allow_origin_regex=r"https://([a-z0-9-]+\.)*vercel\.app$",
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

def _init_supabase() -> Client | None:
    """Connect to Supabase if credentials are valid; never crash app startup."""
    placeholders = {"", "your-anon-key", "your-project.supabase.co"}
    if SUPABASE_URL in placeholders or SUPABASE_ANON_KEY in placeholders:
        logger.warning("Supabase not configured — valuations will not be cached")
        return None
    if not SUPABASE_URL.startswith("https://") or "supabase.co" not in SUPABASE_URL:
        logger.warning("SUPABASE_URL looks invalid — skipping Supabase")
        return None
    if len(SUPABASE_ANON_KEY) < 20:
        logger.warning("SUPABASE_ANON_KEY looks invalid — skipping Supabase")
        return None
    try:
        return create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
    except Exception as exc:
        logger.warning("Supabase client failed: %s — continuing without cache", exc)
        return None


supabase: Client | None = _init_supabase()

anthropic_client = Anthropic(api_key=ANTHROPIC_API_KEY) if ANTHROPIC_API_KEY else None


def _read_text(path: Path) -> str:
    if path.exists():
        return path.read_text(encoding="utf-8")
    return ""


SYSTEM_PROMPT = "\n\n".join(
    filter(None, [_read_text(AGENT_PATH), _read_text(SKILL_PATH)])
)

RESPONSE_SCHEMA = """
Return a single JSON object with this shape:
{
  "ticker": "AAPL",
  "company_name": "Apple Inc.",
  "currency": "USD",
  "current_price": 0.0,
  "market_cap_bn": 0.0,
  "enterprise_value_bn": 0.0,
  "data_sources": ["FMP"],
  "summary": "2-3 sentence investment summary",
  "wacc": {
    "risk_free_rate": 0.0,
    "beta": 0.0,
    "equity_risk_premium": 0.0,
    "cost_of_equity": 0.0,
    "cost_of_debt": 0.0,
    "tax_rate": 0.0,
    "debt_weight": 0.0,
    "equity_weight": 0.0,
    "wacc": 0.0
  },
  "scenarios": {
    "bear": {"implied_price": 0.0, "upside_pct": 0.0, "revenue_cagr": 0.0, "terminal_growth": 0.0, "exit_multiple": 0.0},
    "base": {"implied_price": 0.0, "upside_pct": 0.0, "revenue_cagr": 0.0, "terminal_growth": 0.0, "exit_multiple": 0.0},
    "bull": {"implied_price": 0.0, "upside_pct": 0.0, "revenue_cagr": 0.0, "terminal_growth": 0.0, "exit_multiple": 0.0}
  },
  "projections": {
    "base": {
      "years": ["2025E","2026E","2027E","2028E","2029E"],
      "revenue": [],
      "ebitda": [],
      "ebit": [],
      "fcf": [],
      "revenue_growth": []
    }
  },
  "comps": [
    {"ticker": "MSFT", "ev_ebitda": 0.0, "pe": 0.0, "revenue_growth": 0.0, "ebitda_margin": 0.0}
  ],
  "football_field": [
    {"method": "DCF Bear", "low": 0.0, "high": 0.0, "mid": 0.0},
    {"method": "DCF Base", "low": 0.0, "high": 0.0, "mid": 0.0},
    {"method": "DCF Bull", "low": 0.0, "high": 0.0, "mid": 0.0},
    {"method": "EV/EBITDA Comps", "low": 0.0, "high": 0.0, "mid": 0.0},
    {"method": "P/E Comps", "low": 0.0, "high": 0.0, "mid": 0.0}
  ],
  "historical": {
    "revenue": [],
    "years": [],
    "ebitda_margin": [],
    "fcf_margin": []
  }
}
"""

TOOLS = [
    {
        "name": "get_company_profile",
        "description": "Company profile: price, market cap, sector, beta, shares outstanding.",
        "input_schema": {
            "type": "object",
            "properties": {"ticker": {"type": "string"}},
            "required": ["ticker"],
        },
    },
    {
        "name": "get_income_statement",
        "description": "Annual income statements (last 5 years).",
        "input_schema": {
            "type": "object",
            "properties": {"ticker": {"type": "string"}, "limit": {"type": "integer"}},
            "required": ["ticker"],
        },
    },
    {
        "name": "get_balance_sheet",
        "description": "Annual balance sheets (last 5 years).",
        "input_schema": {
            "type": "object",
            "properties": {"ticker": {"type": "string"}, "limit": {"type": "integer"}},
            "required": ["ticker"],
        },
    },
    {
        "name": "get_cash_flow",
        "description": "Annual cash flow statements (last 5 years).",
        "input_schema": {
            "type": "object",
            "properties": {"ticker": {"type": "string"}, "limit": {"type": "integer"}},
            "required": ["ticker"],
        },
    },
    {
        "name": "get_analyst_estimates",
        "description": "Analyst revenue and EPS estimates.",
        "input_schema": {
            "type": "object",
            "properties": {"ticker": {"type": "string"}},
            "required": ["ticker"],
        },
    },
    {
        "name": "get_peer_companies",
        "description": "List of peer tickers for comps.",
        "input_schema": {
            "type": "object",
            "properties": {"ticker": {"type": "string"}},
            "required": ["ticker"],
        },
    },
    {
        "name": "get_key_metrics",
        "description": "Key valuation metrics for a ticker (use for each peer).",
        "input_schema": {
            "type": "object",
            "properties": {"ticker": {"type": "string"}, "limit": {"type": "integer"}},
            "required": ["ticker"],
        },
    },
]


async def fmp_get(path: str, params: dict[str, Any] | None = None) -> Any:
    if not FMP_API_KEY:
        raise HTTPException(status_code=500, detail="FMP_API_KEY is not configured")
    query = dict(params or {})
    query["apikey"] = FMP_API_KEY
    url = f"{FMP_BASE}/{path.lstrip('/')}"
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.get(url, params=query)
        response.raise_for_status()
        return response.json()


async def run_tool(name: str, inputs: dict[str, Any]) -> Any:
    ticker = inputs.get("ticker", "").upper()
    limit = int(inputs.get("limit", 5))

    if name == "get_company_profile":
        data = await fmp_get(f"profile/{ticker}")
        return data[0] if isinstance(data, list) and data else data

    if name == "get_income_statement":
        return await fmp_get(f"income-statement/{ticker}", {"limit": limit})

    if name == "get_balance_sheet":
        return await fmp_get(f"balance-sheet-statement/{ticker}", {"limit": limit})

    if name == "get_cash_flow":
        return await fmp_get(f"cash-flow-statement/{ticker}", {"limit": limit})

    if name == "get_analyst_estimates":
        return await fmp_get(f"analyst-estimates/{ticker}", {"limit": 5})

    if name == "get_peer_companies":
        peers = await fmp_get(f"stock_peers", {"symbol": ticker})
        if isinstance(peers, list) and peers:
            return peers[0].get("peersList", peers[0])
        return peers

    if name == "get_key_metrics":
        return await fmp_get(f"key-metrics/{ticker}", {"limit": limit})

    raise ValueError(f"Unknown tool: {name}")


def _parse_json_from_text(text: str) -> dict[str, Any]:
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1:
        raise ValueError("Model did not return JSON")
    return json.loads(text[start : end + 1])


async def run_valuation_agent(ticker: str) -> dict[str, Any]:
    if not anthropic_client:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY is not configured")

    messages: list[dict[str, Any]] = [
        {
            "role": "user",
            "content": (
                f"Build a full DCF valuation for {ticker.upper()}. "
                f"Fetch all required data via tools, then respond with JSON only.\n\n"
                f"{RESPONSE_SCHEMA}"
            ),
        }
    ]

    for _ in range(24):
        response = anthropic_client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=16000,
            system=SYSTEM_PROMPT or "You are a valuation analyst.",
            tools=TOOLS,
            messages=messages,
        )

        assistant_content: list[dict[str, Any]] = []
        tool_uses = []

        for block in response.content:
            if block.type == "text":
                assistant_content.append({"type": "text", "text": block.text})
            elif block.type == "tool_use":
                tool_uses.append(block)
                assistant_content.append(
                    {
                        "type": "tool_use",
                        "id": block.id,
                        "name": block.name,
                        "input": block.input,
                    }
                )

        messages.append({"role": "assistant", "content": assistant_content})

        if response.stop_reason == "end_turn" and not tool_uses:
            for block in response.content:
                if block.type == "text":
                    return _parse_json_from_text(block.text)
            raise HTTPException(status_code=502, detail="Agent ended without JSON output")

        if not tool_uses:
            for block in response.content:
                if block.type == "text":
                    try:
                        return _parse_json_from_text(block.text)
                    except ValueError:
                        pass
            break

        tool_results = []
        for tool_use in tool_uses:
            try:
                result = await run_tool(tool_use.name, tool_use.input)
                payload = json.dumps(result)[:120000]
            except Exception as exc:
                payload = json.dumps({"error": str(exc)})
            tool_results.append(
                {
                    "type": "tool_result",
                    "tool_use_id": tool_use.id,
                    "content": payload,
                }
            )

        messages.append({"role": "user", "content": tool_results})

    raise HTTPException(status_code=502, detail="Valuation agent did not complete in time")


def save_to_supabase(ticker: str, payload: dict[str, Any]) -> str | None:
    if not supabase:
        return None
    row = supabase.table("valuations").insert({"ticker": ticker.upper(), "payload": payload}).execute()
    valuation_id = row.data[0]["id"] if row.data else None

    scenarios = payload.get("scenarios", {})
    for name, data in scenarios.items():
        if isinstance(data, dict):
            supabase.table("assumptions").insert(
                {"valuation_id": valuation_id, "scenario": name, "payload": data}
            ).execute()

    base_proj = payload.get("projections", {}).get("base", {})
    if base_proj and valuation_id:
        supabase.table("projections").insert(
            {
                "valuation_id": valuation_id,
                "scenario": "base",
                "year_label": "all",
                "payload": base_proj,
            }
        ).execute()

    for comp in payload.get("comps", []):
        if valuation_id and isinstance(comp, dict):
            supabase.table("comps").insert(
                {
                    "valuation_id": valuation_id,
                    "peer_ticker": comp.get("ticker", ""),
                    "payload": comp,
                }
            ).execute()

    return valuation_id


@app.get("/health")
async def health():
    return {
        "status": "ok",
        "anthropic": bool(ANTHROPIC_API_KEY),
        "fmp": bool(FMP_API_KEY),
        "supabase": bool(supabase),
    }


@app.get("/api/valuation/{ticker}")
async def get_valuation(ticker: str, refresh: bool = False):
    ticker = ticker.upper()

    if not refresh and supabase:
        cached = (
            supabase.table("valuations")
            .select("payload, created_at")
            .eq("ticker", ticker)
            .order("created_at", desc=True)
            .limit(1)
            .execute()
        )
        if cached.data:
            return {"source": "cache", **cached.data[0]["payload"]}

    result = await run_valuation_agent(ticker)
    result["ticker"] = ticker
    result["generated_at"] = datetime.utcnow().isoformat() + "Z"
    save_to_supabase(ticker, result)
    return {"source": "live", **result}


@app.post("/api/export/{ticker}")
async def export_valuation(
    ticker: str,
    body: dict[str, Any] = Body(default_factory=dict),
):
    ticker = ticker.upper()
    data = body or {}
    if not data.get("scenarios"):
        data = await run_valuation_agent(ticker)

    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Summary"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    accent_fill = PatternFill("solid", fgColor="059669")

    ws["A1"] = f"{data.get('company_name', ticker)} — DCF Export"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:F1")

    row = 3
    ws.cell(row, 1, "Scenario").font = header_font
    ws.cell(row, 2, "Implied Price").font = header_font
    ws.cell(row, 3, "Upside %").font = header_font
    for col in range(1, 4):
        ws.cell(row, col).fill = accent_fill
        ws.cell(row, col).font = Font(color="FFFFFF", bold=True)

    scenarios = data.get("scenarios", {})
    for scenario in ("bear", "base", "bull"):
        s = scenarios.get(scenario, {})
        row += 1
        ws.cell(row, 1, scenario.title())
        ws.cell(row, 2, s.get("implied_price"))
        ws.cell(row, 3, s.get("upside_pct"))

    row += 2
    ws.cell(row, 1, "WACC").font = Font(bold=True)
    wacc = data.get("wacc", {})
    for key, value in wacc.items():
        row += 1
        ws.cell(row, 1, key)
        ws.cell(row, 2, value)

    proj = data.get("projections", {}).get("base", {})
    if proj:
        row += 2
        ws.cell(row, 1, "Base projections").font = Font(bold=True)
        row += 1
        years = proj.get("years", [])
        for i, year in enumerate(years, start=2):
            ws.cell(row, i, year)
        for label, key in [("Revenue", "revenue"), ("EBITDA", "ebitda"), ("FCF", "fcf")]:
            row += 1
            ws.cell(row, 1, label)
            values = proj.get(key, [])
            for i, val in enumerate(values, start=2):
                ws.cell(row, i, val)

    for col_idx in range(1, (ws.max_column or 1) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, (ws.max_row or 1) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 24)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{ticker}_valuation.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
